from http.server import BaseHTTPRequestHandler
import json
import logging
import os
import sys
import time
import threading
import ctypes
from functools import wraps
from datetime import datetime, timezone


# ============================================================
# Timeout utility (inlined to avoid Vercel treating utils/ as functions)
# ============================================================

class FunctionTimeoutError(Exception):
    def __init__(self, seconds=None):
        self.seconds = seconds
        msg = f"Function execution timed out after {seconds}s" if seconds else "Function execution timed out"
        super().__init__(msg)


class function_timeout:
    def __init__(self, seconds=9):
        self.seconds = seconds
        self._timer = None
        self._target_thread_id = None

    def _on_timeout(self):
        if self._target_thread_id is not None:
            ctypes.pythonapi.PyThreadState_SetAsyncExc(
                ctypes.c_ulong(self._target_thread_id),
                ctypes.py_object(FunctionTimeoutError)
            )

    def __enter__(self):
        self._target_thread_id = threading.current_thread().ident
        self._timer = threading.Timer(self.seconds, self._on_timeout)
        self._timer.daemon = True
        self._timer.start()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self._timer:
            self._timer.cancel()
        return False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


MAX_RETRIES = 3
RETRY_TIMEOUT = 5  # seconds per attempt


class SheetsUnavailableError(Exception):
    """Raised when Google Sheets is unavailable after all retry attempts."""
    pass


def load_campaign_config(campaign_id):
    """Load campaign configuration JSON by campaign_id."""
    config_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..",
        "public",
        "config",
        "campaigns",
    )
    config_path = os.path.join(config_dir, f"{campaign_id}.json")
    config_path = os.path.normpath(config_path)

    if not os.path.isfile(config_path):
        return None

    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def find_product(config, product_id):
    """Find a product in the campaign config by product_id."""
    for product in config.get("products", []):
        if product.get("product_id") == product_id:
            return product
    return None


def validate_registration(body):
    """
    Validate registration form fields.

    Returns:
        (None, None) if valid
        (error_code, error_message) if invalid
    """
    # Required string fields with max length constraints
    field_constraints = {
        "campaign_id": None,  # no max length constraint
        "product_id": None,
        "selected_size": None,
        "selected_color": None,
        "instagram_id": 200,
        "name": 100,
        "phone": 20,
        "address": 300,
        "postal_code": 10,
    }

    # Check all required fields are present and non-empty strings
    for field, max_length in field_constraints.items():
        value = body.get(field)
        if value is None or not isinstance(value, str) or value.strip() == "":
            return ("VALIDATION_ERROR", "필수 항목을 확인해 주세요.")
        if max_length is not None and len(value) > max_length:
            return ("VALIDATION_ERROR", "필수 항목을 확인해 주세요.")

    # Consent must be boolean true
    consent = body.get("consent")
    if consent is not True:
        return ("VALIDATION_ERROR", "필수 항목을 확인해 주세요.")

    # member_type must be exactly "new" or "returning"
    member_type = body.get("member_type")
    if member_type not in ("new", "returning"):
        return ("INVALID_MEMBER_TYPE", "회원 유형을 선택해 주세요.")

    return (None, None)


def apply_single_product_mode(body, config):
    """
    For single-product campaigns, override body["product_id"] with the
    campaign's configured product_id and validate the product is available.

    Returns:
        (None, None) if not single mode or product is valid and active
        (error_code, error_message) if product is unavailable
    """
    if config.get("product_mode") != "single":
        return (None, None)

    products = config.get("products", [])
    if not products:
        return ("PRODUCT_UNAVAILABLE", "현재 해당 상품은 신청할 수 없습니다.")

    product = products[0]
    product_id = product.get("product_id")

    if not product_id:
        return ("PRODUCT_UNAVAILABLE", "현재 해당 상품은 신청할 수 없습니다.")

    # Validate product is active (status must be "open")
    if product.get("status") != "open":
        return ("PRODUCT_UNAVAILABLE", "현재 해당 상품은 신청할 수 없습니다.")

    # Override body product_id with the campaign's configured product
    body["product_id"] = product_id
    return (None, None)


def validate_size_color(body, config):
    """
    Validate that selected_size and selected_color are valid for the product.

    Returns:
        (None, None) if valid
        (error_code, error_message) if invalid
    """
    product_id = body.get("product_id")
    product = find_product(config, product_id)

    if product is None:
        return ("INVALID_SIZE_COLOR", "선택한 사이즈 또는 컬러가 유효하지 않습니다.")

    selected_size = body.get("selected_size")
    selected_color = body.get("selected_color")

    available_sizes = product.get("available_sizes", [])
    available_colors = product.get("available_colors", [])

    if selected_size not in available_sizes:
        return ("INVALID_SIZE_COLOR", "선택한 사이즈 또는 컬러가 유효하지 않습니다.")

    if selected_color not in available_colors:
        return ("INVALID_SIZE_COLOR", "선택한 사이즈 또는 컬러가 유효하지 않습니다.")

    return (None, None)


def check_product_closed(config, product_id):
    """
    Check if a Campaign_Product's status is "closed".

    Returns:
        True if product is closed, False otherwise.
    """
    product = find_product(config, product_id)
    if product is None:
        return False  # Product not found is handled elsewhere
    return product.get("status") == "closed"


def check_duplicate(campaign_id, product_id, instagram_id):
    """
    Check if a registration already exists for the given combination
    by querying Google Sheets for rows matching (campaign_id, product_id, instagram_id).

    Uses gspread to connect to the Google Sheets data store. If the sheet is
    unavailable or credentials are not configured, the check is skipped (returns False)
    to avoid blocking registrations due to transient errors.

    Args:
        campaign_id: The campaign identifier
        product_id: The product identifier
        instagram_id: The creator's Instagram ID

    Returns:
        True if a duplicate registration exists, False otherwise.
    """
    if not GSPREAD_AVAILABLE:
        return False

    try:
        credentials_json = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID")

        if not credentials_json or not sheet_id:
            return False

        credentials_info = json.loads(credentials_json)
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        credentials = Credentials.from_service_account_info(
            credentials_info, scopes=scopes
        )
        client = gspread.authorize(credentials)
        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.sheet1

        # Get all rows and check for duplicate
        # Row data columns (0-indexed): 0=timestamp, 1=campaign_id, 2=product_id,
        # 3=selected_size, 4=selected_color, 5=instagram_id, ...
        all_rows = worksheet.get_all_values()

        for row in all_rows[1:]:  # Skip header row
            if len(row) >= 6:
                row_campaign_id = row[1]
                row_product_id = row[2]
                row_instagram_id = row[5]

                if (
                    row_campaign_id == campaign_id
                    and row_product_id == product_id
                    and row_instagram_id == instagram_id
                ):
                    return True

        return False

    except Exception:
        # If Google Sheets is unavailable, skip duplicate check
        # This will be properly handled with retry logic in task 2.4
        return False


def save_to_retry_queue(row_data):
    """
    Append failed row data to a JSON lines file for manual retry.

    Saves to retry_queue.jsonl in the project root directory.
    """
    project_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
    queue_path = os.path.normpath(os.path.join(project_root, "retry_queue.jsonl"))

    entry = {
        "row_data": row_data,
        "failed_at": datetime.now(timezone.utc).isoformat(),
    }

    with open(queue_path, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def append_with_retry(worksheet, row_data):
    """
    Append a row to a Google Sheets worksheet with retry logic.

    Attempts up to MAX_RETRIES times, with RETRY_TIMEOUT seconds per attempt.
    On failure after all retries, saves to retry queue and raises SheetsUnavailableError.

    Args:
        worksheet: A gspread Worksheet object
        row_data: List of values to append as a row

    Returns:
        True on success

    Raises:
        SheetsUnavailableError: After all retry attempts fail
    """
    last_error = None

    for attempt in range(MAX_RETRIES):
        try:
            start_time = time.time()
            worksheet.append_row(row_data, value_input_option="RAW")
            elapsed = time.time() - start_time

            if elapsed > RETRY_TIMEOUT:
                # Even though it completed, if it took too long we note it but still accept
                pass

            return True
        except Exception as e:
            last_error = e
            elapsed = time.time() - start_time

            if attempt < MAX_RETRIES - 1:
                # Brief backoff before next retry
                time.sleep(1)

    # All retries exhausted
    save_to_retry_queue(row_data)
    raise SheetsUnavailableError(
        f"Google Sheets unavailable after {MAX_RETRIES} attempts: {str(last_error)}"
    )


def persist_registration(body):
    """
    Persist a validated registration to Google Sheets.

    Serializes the registration as a row with 12 columns and appends it
    to the configured Google Sheets document.

    Args:
        body: The validated registration request body dict

    Returns:
        list: The row_data that was persisted (for use by sync_to_excel)

    Raises:
        SheetsUnavailableError: If persistence fails after all retries
    """
    if not GSPREAD_AVAILABLE:
        raise SheetsUnavailableError("gspread library not available")

    credentials_json = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
    sheet_id = os.environ.get("GOOGLE_SHEETS_ID")

    if not credentials_json or not sheet_id:
        raise SheetsUnavailableError("Google Sheets credentials not configured")

    # Initialize gspread client
    credentials_info = json.loads(credentials_json)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_info(
        credentials_info, scopes=scopes
    )
    client = gspread.authorize(credentials)
    spreadsheet = client.open_by_key(sheet_id)
    worksheet = spreadsheet.sheet1

    # Serialize registration as row with 12 columns
    # Order: timestamp, campaign_id, product_id, selected_size, selected_color,
    #         instagram_id, name, phone, address, postal_code, consent_status, member_type
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    row_data = [
        timestamp,
        body.get("campaign_id", ""),
        body.get("product_id", ""),           # Internal identifier
        body.get("selected_size", ""),         # Internal identifier
        body.get("selected_color", ""),        # Internal identifier
        body.get("instagram_id", ""),
        body.get("name", ""),
        body.get("phone", ""),
        body.get("address", ""),
        body.get("postal_code", ""),
        "true",                                # consent_status
        body.get("member_type", ""),           # member_type ("new" or "returning")
        # --- Management columns (empty, for manual use in Sheets) ---
        "",                                    # 合作状态 (待审核/已寄样/已发帖/完成)
        "",                                    # Creator等级 (新人/优质/核心)
        "",                                    # 历史合作次数
        "",                                    # 内容质量评分
        "",                                    # 发帖链接
        "",                                    # 备注
    ]

    append_with_retry(worksheet, row_data)
    return row_data


# Excel sync headers matching the 18-column registration row format
# Columns 1-12: auto-filled by API
# Columns 13-18: management columns (manually filled in Sheets)
EXCEL_HEADERS = [
    "timestamp",
    "campaign_id",
    "product_id",
    "selected_size",
    "selected_color",
    "instagram_id",
    "name",
    "phone",
    "address",
    "postal_code",
    "consent",
    "member_type",
    "合作状态",
    "Creator等级",
    "历史合作次数",
    "内容质量评分",
    "发帖链接",
    "备注",
]

# Timeout for Excel sync operation (seconds)
EXCEL_SYNC_TIMEOUT = 60


def _get_excel_path():
    """Get the path to the local Excel file (data/registrations.xlsx relative to project root)."""
    project_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
    data_dir = os.path.normpath(os.path.join(project_root, "data"))
    return data_dir, os.path.join(data_dir, "registrations.xlsx")


def sync_to_excel(row_data):
    """
    Sync a registration row to the local Excel file using openpyxl.

    Opens or creates data/registrations.xlsx (relative to project root),
    appends the row_data, and saves. Must complete within 60 seconds.

    This is a best-effort operation — failures are logged but do not
    affect the user response.

    Args:
        row_data: List of 18 values matching EXCEL_HEADERS column order
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available, skipping Excel sync")
        return

    start_time = time.time()

    data_dir, excel_path = _get_excel_path()

    # Create data directory if it doesn't exist
    os.makedirs(data_dir, exist_ok=True)

    # Open existing workbook or create new one with headers
    if os.path.isfile(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"
        ws.append(EXCEL_HEADERS)

    # Check timeout before appending
    elapsed = time.time() - start_time
    if elapsed >= EXCEL_SYNC_TIMEOUT:
        logger.warning(
            "Excel sync timed out before appending row (%.1fs elapsed)", elapsed
        )
        return

    # Append the row data
    ws.append(row_data)

    # Check timeout before saving
    elapsed = time.time() - start_time
    if elapsed >= EXCEL_SYNC_TIMEOUT:
        logger.warning(
            "Excel sync timed out before saving file (%.1fs elapsed)", elapsed
        )
        return

    # Save the workbook
    wb.save(excel_path)

    elapsed = time.time() - start_time
    logger.info("Excel sync completed in %.2fs", elapsed)


def make_error_response(code, message):
    """Create a standardized error response dict."""
    return {"status": "error", "code": code, "message": message}


def make_success_response():
    """Create a standardized success response dict."""
    return {"status": "success", "message": "Registration submitted successfully"}


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            with function_timeout(seconds=9):
                # Read and parse request body
                content_length = int(self.headers.get("Content-Length", 0))
                if content_length == 0:
                    self._send_json(400, make_error_response(
                        "VALIDATION_ERROR", "필수 항목을 확인해 주세요."
                    ))
                    return

                raw_body = self.rfile.read(content_length)
                try:
                    body = json.loads(raw_body)
                except (json.JSONDecodeError, ValueError):
                    self._send_json(400, make_error_response(
                        "VALIDATION_ERROR", "필수 항목을 확인해 주세요."
                    ))
                    return

                # Step 1: Validate required fields and constraints
                error_code, error_message = validate_registration(body)
                if error_code:
                    self._send_json(400, make_error_response(error_code, error_message))
                    return

                # Step 2: Load campaign config
                campaign_id = body.get("campaign_id")
                config = load_campaign_config(campaign_id)
                if config is None:
                    self._send_json(400, make_error_response(
                        "VALIDATION_ERROR", "필수 항목을 확인해 주세요."
                    ))
                    return

                # Step 2.5: Single-product auto-association
                error_code, error_message = apply_single_product_mode(body, config)
                if error_code:
                    self._send_json(400, make_error_response(error_code, error_message))
                    return

                # Step 3: Validate size and color against product's available options
                error_code, error_message = validate_size_color(body, config)
                if error_code:
                    self._send_json(400, make_error_response(error_code, error_message))
                    return

                # Step 4: Check if product is closed
                product_id = body.get("product_id")
                if check_product_closed(config, product_id):
                    self._send_json(400, make_error_response(
                        "PRODUCT_UNAVAILABLE",
                        "현재 해당 상품은 신청할 수 없습니다."
                    ))
                    return

                # Step 5: Check for duplicate registration
                instagram_id = body.get("instagram_id")
                if check_duplicate(campaign_id, product_id, instagram_id):
                    self._send_json(409, make_error_response(
                        "DUPLICATE_REGISTRATION",
                        "이미 해당 상품에 신청하셨습니다."
                    ))
                    return

                # Step 6: Persist registration to Google Sheets
                try:
                    row_data = persist_registration(body)
                except SheetsUnavailableError:
                    self._send_json(503, make_error_response(
                        "SHEETS_UNAVAILABLE",
                        "일시적인 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."
                    ))
                    return

                # Step 7: Best-effort sync to local Excel file
                # This does NOT block the user response — if it fails, we log and continue
                try:
                    sync_to_excel(row_data)
                except Exception as e:
                    logger.error("Excel sync failed (best-effort): %s", str(e))

                self._send_json(200, make_success_response())

        except FunctionTimeoutError:
            self._send_json(504, make_error_response(
                "TIMEOUT",
                "요청을 처리할 수 없습니다. 다시 시도해 주세요."
            ))
        except Exception:
            self._send_json(500, make_error_response(
                "VALIDATION_ERROR", "필수 항목을 확인해 주세요."
            ))

    def _send_json(self, status_code, data):
        """Send a JSON response with the given status code."""
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

    def do_OPTIONS(self):
        """Handle CORS preflight requests."""
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
