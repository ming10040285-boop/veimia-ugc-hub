"""Unit tests for the Excel sync functionality (Task 2.5)."""
import json
import os
import sys
import tempfile
import time
import pytest
from unittest.mock import patch, MagicMock

# Add the project root to path so we can import from api
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from api.register import (
    sync_to_excel,
    _get_excel_path,
    EXCEL_HEADERS,
    EXCEL_SYNC_TIMEOUT,
)


def make_sample_row_data():
    """Create sample row data matching the 11-column registration format."""
    return [
        "2024-01-15T10:30:00Z",  # timestamp
        "campaign-001",           # campaign_id
        "prod-001",               # product_id
        "M",                      # selected_size
        "블랙",                    # selected_color
        "@creator_test",          # instagram_id
        "홍길동",                   # name
        "010-1234-5678",          # phone
        "서울시 강남구 테헤란로 123",  # address
        "06234",                  # postal_code
        "true",                   # consent
    ]


class TestSyncToExcel:
    """Tests for sync_to_excel function."""

    def test_creates_new_file_with_headers(self, tmp_path):
        """When no Excel file exists, sync_to_excel creates one with headers."""
        data_dir = tmp_path / "data"
        excel_path = str(data_dir / "registrations.xlsx")

        with patch("api.register._get_excel_path", return_value=(str(data_dir), excel_path)):
            row_data = make_sample_row_data()
            sync_to_excel(row_data)

        # Verify file was created
        assert os.path.isfile(excel_path)

        # Verify contents
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        # Check headers in row 1
        headers = [cell.value for cell in ws[1]]
        assert headers == EXCEL_HEADERS

        # Check data in row 2
        data_row = [cell.value for cell in ws[2]]
        assert data_row == row_data

    def test_appends_to_existing_file(self, tmp_path):
        """When Excel file exists, sync_to_excel appends to it."""
        from openpyxl import Workbook

        data_dir = tmp_path / "data"
        data_dir.mkdir()
        excel_path = str(data_dir / "registrations.xlsx")

        # Create existing file with header and one row
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADERS)
        ws.append(["2024-01-14T09:00:00Z", "camp-old", "prod-old", "S", "화이트",
                   "@old_creator", "김철수", "010-0000-0000", "부산", "12345", "true"])
        wb.save(excel_path)

        with patch("api.register._get_excel_path", return_value=(str(data_dir), excel_path)):
            row_data = make_sample_row_data()
            sync_to_excel(row_data)

        # Verify new row was appended
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        assert ws.max_row == 3  # header + old row + new row

        # Verify the new row is in position 3
        new_row = [cell.value for cell in ws[3]]
        assert new_row == row_data

    def test_creates_data_directory_if_missing(self, tmp_path):
        """sync_to_excel creates the data/ directory if it doesn't exist."""
        data_dir = str(tmp_path / "nonexistent" / "data")
        excel_path = os.path.join(data_dir, "registrations.xlsx")

        with patch("api.register._get_excel_path", return_value=(data_dir, excel_path)):
            row_data = make_sample_row_data()
            sync_to_excel(row_data)

        assert os.path.isdir(data_dir)
        assert os.path.isfile(excel_path)

    def test_preserves_existing_data(self, tmp_path):
        """sync_to_excel does not modify existing rows."""
        from openpyxl import Workbook, load_workbook

        data_dir = tmp_path / "data"
        data_dir.mkdir()
        excel_path = str(data_dir / "registrations.xlsx")

        # Create existing file with multiple rows
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADERS)
        existing_rows = [
            ["2024-01-10T08:00:00Z", "camp-1", "prod-1", "S", "핑크",
             "@creator1", "이영희", "010-1111-1111", "대전", "11111", "true"],
            ["2024-01-11T09:00:00Z", "camp-1", "prod-2", "L", "블랙",
             "@creator2", "박민수", "010-2222-2222", "인천", "22222", "true"],
        ]
        for row in existing_rows:
            ws.append(row)
        wb.save(excel_path)

        with patch("api.register._get_excel_path", return_value=(str(data_dir), excel_path)):
            sync_to_excel(make_sample_row_data())

        # Verify existing rows are untouched
        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 4  # header + 2 existing + 1 new

        row2 = [cell.value for cell in ws[2]]
        assert row2 == existing_rows[0]

        row3 = [cell.value for cell in ws[3]]
        assert row3 == existing_rows[1]

    def test_handles_korean_characters(self, tmp_path):
        """sync_to_excel correctly stores Korean text."""
        data_dir = tmp_path / "data"
        excel_path = str(data_dir / "registrations.xlsx")

        with patch("api.register._get_excel_path", return_value=(str(data_dir), excel_path)):
            row_data = make_sample_row_data()
            sync_to_excel(row_data)

        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        data_row = [cell.value for cell in ws[2]]
        # Check Korean fields
        assert data_row[4] == "블랙"          # selected_color
        assert data_row[6] == "홍길동"         # name
        assert data_row[8] == "서울시 강남구 테헤란로 123"  # address

    def test_skips_when_openpyxl_unavailable(self):
        """sync_to_excel skips gracefully when openpyxl is not available."""
        with patch("api.register.OPENPYXL_AVAILABLE", False):
            # Should not raise, just return
            sync_to_excel(make_sample_row_data())

    def test_worksheet_title_is_registrations(self, tmp_path):
        """New workbooks should have the sheet titled 'Registrations'."""
        data_dir = tmp_path / "data"
        excel_path = str(data_dir / "registrations.xlsx")

        with patch("api.register._get_excel_path", return_value=(str(data_dir), excel_path)):
            sync_to_excel(make_sample_row_data())

        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        assert wb.active.title == "Registrations"


class TestExcelSyncTimeout:
    """Tests for the 60-second timeout constraint."""

    def test_timeout_constant_is_60(self):
        """EXCEL_SYNC_TIMEOUT should be 60 seconds."""
        assert EXCEL_SYNC_TIMEOUT == 60


class TestGetExcelPath:
    """Tests for _get_excel_path helper."""

    def test_returns_correct_structure(self):
        """_get_excel_path returns (data_dir, excel_path) tuple."""
        data_dir, excel_path = _get_excel_path()
        assert data_dir.endswith("data")
        assert excel_path.endswith("registrations.xlsx")
        assert excel_path.startswith(data_dir)


class TestExcelSyncIntegration:
    """Integration tests for Excel sync in the registration flow."""

    @patch("api.register.check_duplicate", return_value=False)
    @patch("api.register.persist_registration")
    @patch("api.register.sync_to_excel")
    @patch("api.register.load_campaign_config")
    def test_sync_called_after_successful_persist(
        self, mock_load_config, mock_sync, mock_persist, mock_dup
    ):
        """sync_to_excel is called after persist_registration succeeds."""
        from io import BytesIO
        from api.register import handler
        from unittest.mock import MagicMock

        mock_load_config.return_value = {
            "campaign_id": "sample",
            "products": [
                {
                    "product_id": "prod-001",
                    "available_sizes": ["S", "M", "L", "XL", "2XL"],
                    "available_colors": ["아이보리", "블랙", "핑크", "라벤더", "스킨베이지"],
                    "status": "open",
                }
            ],
        }
        row_data = make_sample_row_data()
        mock_persist.return_value = row_data

        body = {
            "campaign_id": "sample",
            "product_id": "prod-001",
            "selected_size": "M",
            "selected_color": "블랙",
            "instagram_id": "@creator_test",
            "name": "홍길동",
            "phone": "010-1234-5678",
            "address": "서울시 강남구 테헤란로 123",
            "postal_code": "06234",
            "consent": True,
        }
        body_bytes = json.dumps(body).encode("utf-8")

        mock_handler = MagicMock(spec=handler)
        mock_handler.headers = {"Content-Length": str(len(body_bytes))}
        mock_handler.rfile = BytesIO(body_bytes)

        responses = []

        def mock_send_json(status_code, data):
            responses.append((status_code, data))

        mock_handler._send_json = mock_send_json

        handler.do_POST(mock_handler)

        # Verify sync was called with the row data
        mock_sync.assert_called_once_with(row_data)
        # Verify success response still returned
        assert responses[0][0] == 200

    @patch("api.register.check_duplicate", return_value=False)
    @patch("api.register.persist_registration")
    @patch("api.register.sync_to_excel")
    @patch("api.register.load_campaign_config")
    def test_sync_failure_does_not_affect_response(
        self, mock_load_config, mock_sync, mock_persist, mock_dup
    ):
        """If sync_to_excel raises an exception, user still gets success response."""
        from io import BytesIO
        from api.register import handler
        from unittest.mock import MagicMock

        mock_load_config.return_value = {
            "campaign_id": "sample",
            "products": [
                {
                    "product_id": "prod-001",
                    "available_sizes": ["S", "M", "L", "XL", "2XL"],
                    "available_colors": ["아이보리", "블랙", "핑크", "라벤더", "스킨베이지"],
                    "status": "open",
                }
            ],
        }
        row_data = make_sample_row_data()
        mock_persist.return_value = row_data
        mock_sync.side_effect = Exception("Disk full")

        body = {
            "campaign_id": "sample",
            "product_id": "prod-001",
            "selected_size": "M",
            "selected_color": "블랙",
            "instagram_id": "@creator_test",
            "name": "홍길동",
            "phone": "010-1234-5678",
            "address": "서울시 강남구 테헤란로 123",
            "postal_code": "06234",
            "consent": True,
        }
        body_bytes = json.dumps(body).encode("utf-8")

        mock_handler = MagicMock(spec=handler)
        mock_handler.headers = {"Content-Length": str(len(body_bytes))}
        mock_handler.rfile = BytesIO(body_bytes)

        responses = []

        def mock_send_json(status_code, data):
            responses.append((status_code, data))

        mock_handler._send_json = mock_send_json

        handler.do_POST(mock_handler)

        # User still gets success even though Excel sync failed
        assert responses[0][0] == 200
        assert responses[0][1]["status"] == "success"
